"""Ведомость НЕДОСТАЮЩИХ данных и документов (v0.36).

Жалоба пользователя на итоговое заключение: «в большинстве ответов отсутствуют
исходные данные» — слишком общо, «надо точнее: чего не хватает, в каких
разделах и что необходимо дать».

Здесь собирается КОНКРЕТНЫЙ реестр: по каждому замечанию — чего не хватает,
к какому тому/разделу это относится, что нужно получить и от кого. Сводка
группируется по типу (изыскания, протоколы, договоры, справки, расчёты) и
выгружается в xlsx/docx для рассылки исполнителям.
"""
from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from ..paths import project_paths

# тип нехватки → как это назвать в ведомости (порядок = приоритет определения)
KINDS: list[tuple[str, str, str]] = [
    ("изыскания", r"изыскан|иэи|игми|игди|игэи|инженерно-эколог|инженерно-геолог",
     "Результаты изысканий"),
    ("протоколы", r"протокол|лаборатор|замер|испытан|исследован проб|анализ проб",
     "Протоколы измерений/испытаний"),
    ("договоры", r"договор|соглашен|контракт|полигон|лиценз|специализированн организац",
     "Договоры и лицензии"),
    ("справки", r"справк|письм|запрос|ответ .*орган|цгмс|росгидромет|роспотребнадзор|"
                r"минсельхоз|лесничеств|кадастр|выписк|согласован",
     "Справки и согласования уполномоченных органов"),
    ("расчёты", r"расчёт|расчет|перерасч|рассеиван|упрза|акустическ|шум|"
                r"обоснован объём|таблиц выброс",
     "Расчёты и обоснования"),
    ("иное", r".", "Иные данные"),
]


def _kind(text: str) -> tuple[str, str]:
    t = (text or "").lower()
    for key, pat, human in KINDS:
        if re.search(pat, t):
            return key, human
    return "иное", "Иные данные"


def collect_gaps(project: str) -> dict[str, Any]:
    """Собрать реестр пробелов из ответов Блока 1."""
    from ..pipeline.block1_answers import load_answers
    rows: list[dict] = []
    for a in (load_answers(project) or {}).get("answers", []):
        need: list[str] = []
        md = (a.get("missing_data") or "").strip()
        if md:
            need.append(md)
        for x in (a.get("attachments") or []):
            x = str(x).strip()
            if x:
                need.append(x)
        # ответ без опоры на источники — это тоже пробел: подтверждать нечем
        if not need and (a.get("low_support") or a.get("sources_unverified")):
            need.append("Нет подтверждающих фрагментов в проектной документации — "
                        "требуется первичный документ/расчёт")
        for item in need:
            key, human = _kind(item)
            rows.append({
                "number": str(a.get("number", "")),
                "kind": key,
                "kind_human": human,
                "need": item,
                "volume": a.get("oos_volume", "") or "",
                "location": a.get("edit_location", "") or "",
                "category": a.get("category", "") or "",
                "remark": (a.get("remark") or "")[:400],
                "status": a.get("status", "proposed"),
            })
    by_kind: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_kind[r["kind_human"]].append(r)
    remarks_with_gaps = sorted({r["number"] for r in rows}, key=lambda s: (len(s), s))
    return {"project": project, "rows": rows, "by_kind": dict(by_kind),
            "remarks_with_gaps": remarks_with_gaps,
            "total_remarks": len((load_answers(project) or {}).get("answers", []))}


def summary_text(project: str, gaps: dict | None = None) -> str:
    """Короткая КОНКРЕТНАЯ сводка для итогового заключения (Блок 3)."""
    g = gaps or collect_gaps(project)
    if not g["rows"]:
        return "Недостающих исходных данных не выявлено."
    parts = [f"Пробелы выявлены по {len(g['remarks_with_gaps'])} замечаниям "
             f"из {g['total_remarks']}. Требуется предоставить:"]
    for human, items in sorted(g["by_kind"].items(), key=lambda kv: -len(kv[1])):
        nums = ", ".join(sorted({i["number"] for i in items},
                                key=lambda s: (len(s), s))[:14])
        vols = sorted({i["volume"] for i in items if i["volume"]})[:3]
        example = items[0]["need"][:160]
        parts.append(f"• {human} — замечания № {nums}"
                     + (f"; тома: {', '.join(vols)}" if vols else "")
                     + f". Например: {example}")
    return "\n".join(parts)


def build_gaps_xlsx(project: str, *, out_path: str | Path | None = None) -> Path:
    """Ведомость недостающих данных в xlsx (по одной строке на пробел)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    g = collect_gaps(project)
    wb = Workbook()
    ws = wb.active
    ws.title = "Недостающие данные"
    headers = ["№ замечания", "Тип нехватки", "Что необходимо предоставить",
               "Том ООС", "Место правки", "Категория замечания", "Статус ответа",
               "Текст замечания (начало)"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="1F4E79")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = head_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ST_RU = {"proposed": "не проверен", "accepted": "принят",
             "edited": "правка", "rejected": "отклонён"}
    for r in sorted(g["rows"], key=lambda x: (x["kind_human"], len(x["number"]), x["number"])):
        ws.append([r["number"], r["kind_human"], r["need"], r["volume"],
                   r["location"], r["category"],
                   ST_RU.get(r["status"], r["status"]), r["remark"]])
    for col, width in zip("ABCDEFGH", (12, 26, 60, 26, 26, 20, 14, 60)):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Сводка")
    ws2.append(["Тип нехватки", "Замечаний", "Позиций"])
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = head_fill
    for human, items in sorted(g["by_kind"].items(), key=lambda kv: -len(kv[1])):
        ws2.append([human, len({i["number"] for i in items}), len(items)])
    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12

    out = Path(out_path) if out_path else (
        project_paths(project)["out"] / "ВЕДОМОСТЬ_недостающих_данных.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out
