"""Сводная ТАБЛИЦА ИЗМЕНЕНИЙ по замечаниям (ТЗ 27.07, п. «ОТВЕТЫ»).

«Важно что на что меняется — нужны чёткие видимые изменения в виде таблицы».
Одна строка = одна правка: № замечания, ГДЕ (том/пункт), БЫЛО (текущий текст),
СТАЛО (новый текст), что приложить. Выгружается в xlsx для работы по тому.
"""
from __future__ import annotations

from pathlib import Path

from ..paths import project_paths


def changes_rows(project: str) -> list[dict]:
    """Строки таблицы изменений из ответов Блока 1 (только где есть правка)."""
    from ..pipeline.block1_answers import load_answers
    rows: list[dict] = []
    for a in (load_answers(project) or {}).get("answers", []):
        if a.get("status") == "rejected":
            continue                      # отклонённое пользователем — не правка (ревью)
        was, shall = (a.get("edit_was") or "").strip(), (a.get("edit_shall") or "").strip()
        corr = (a.get("correction") or "").strip()
        if not (was or shall or corr):
            continue                      # ответ без правки текста — не в таблицу
        rows.append({
            "number": str(a.get("number", "")),
            "location": (a.get("edit_location") or "").strip()
                        or (a.get("oos_volume") or "").strip() or "—",
            "was": was or "—",
            "shall": shall or corr or "—",
            "attachments": "; ".join(a.get("attachments") or []),
            "status": a.get("status", "proposed"),
        })
    rows.sort(key=lambda r: (len(r["number"]), r["number"]))
    return rows


def build_changes_xlsx(project: str, *, out_path: str | Path | None = None) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    rows = changes_rows(project)
    wb = Workbook()
    ws = wb.active
    ws.title = "Таблица изменений"
    headers = ["№ замечания", "ГДЕ (том, пункт, таблица)", "БЫЛО (как сейчас)",
               "СТАЛО (новая редакция)", "Приложить", "Статус"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="1F4E79")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = head_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)
    st_ru = {"proposed": "не проверен", "accepted": "принят",
             "edited": "правка", "rejected": "отклонён"}
    for r in rows:
        ws.append([r["number"], r["location"], r["was"], r["shall"],
                   r["attachments"], st_ru.get(r["status"], r["status"])])
    for col, width in zip("ABCDEF", (12, 30, 52, 52, 32, 12)):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    out = Path(out_path) if out_path else (
        project_paths(project)["out"] / "ТАБЛИЦА_изменений.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out
