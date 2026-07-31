"""Импорт результатов расчёта ИЗ УПРЗА (ТЗ 27.07: «выгрузка/загрузка в/из УПРЗА»).

УПРЗА «Эколог» (ИНТЕГРАЛ) выгружает результаты рассеивания таблицами
(txt/csv/xlsx): вещество (код + название), максимальная концентрация в долях
ПДК, расчётная точка. Точный шаблон у версий разный, поэтому парсер ТЕРПИМЫЙ:
ищет строки «4-значный код вещества + числа», а не жёсткие колонки.

Результат хранится в базе проекта (uprza_results.json) и виден в УПРЗА и в
ответах: превышения (>1 ПДК) — готовый материал для замечаний по рассеиванию.
"""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from ..paths import project_paths

# код ЗВ (0301, 2902…) — отдельное 4-значное слово; имя вещества — русские
# буквы БЕЗ цифр (иначе в txt с пробельными колонками имя «съедало» числа и
# строка терялась — найдено ревью); значения — числа после имени
_CODE_RX = re.compile(r"(?:^|[\s;\t|])(\d{4})(?=[\s;\t|])")
_NAME_RX = re.compile(r"[А-Яа-яЁё][А-Яа-яЁё\s\-(),./]{1,60}")
_NUM_RX = re.compile(r"\d+(?:[.,]\d+)?(?:[eEеЕ][-+]?\d+)?")


def results_path(project: str) -> Path:
    return project_paths(project)["root"] / "uprza_results.json"


def _decode(raw: bytes) -> str:
    for enc in ("utf-8-sig", "cp1251", "cp866"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def parse_uprza_file(path: Path) -> list[dict]:
    """Строки результатов: {code, name, max_pdk, raw}. Терпимо к формату."""
    ext = path.suffix.lower()
    lines: list[str] = []
    if ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [("" if c is None else str(c)).strip() for c in row]
                if any(cells):
                    lines.append(";".join(cells))
        wb.close()
    else:                                     # txt / csv / выгрузка отчёта
        lines = _decode(path.read_bytes()).splitlines()

    rows: list[dict] = []
    for ln in lines:
        m = _CODE_RX.search(" " + ln)          # пробел: код может стоять с 0-й позиции
        if not m:
            continue
        code = m.group(1)
        if 1900 <= int(code) <= 2100:          # год, а не код ЗВ (ревью)
            continue
        tail0 = ln[m.end() - 1:]
        nm = _NAME_RX.search(tail0)
        name = (nm.group(0).strip(" -–—:.") if nm else "")
        tail = tail0[nm.end():] if nm else tail0
        nums: list[float] = []
        for s in _NUM_RX.findall(tail):
            try:
                nums.append(float(s.replace(",", ".").replace("е", "e").replace("Е", "E")))
            except ValueError:
                continue
        # доля ПДК — ПЕРВОЕ дробное число разумного диапазона. Целые (номера
        # точек, высоты) отбрасываются; «первое», а не max — потому что проценты
        # вклада (96,5%) идут в строке ПОСЛЕ концентрации и давали ложные
        # «превышения» (сценарий из ревью: «0,85;25;96,5» → 96,5 «ПДК»)
        pdk_cands = [n for n in nums if 0 < n <= 100 and n != int(n)]
        if not pdk_cands:
            continue
        rows.append({"code": code, "name": name, "max_pdk": pdk_cands[0],
                     "raw": ln.strip()[:200]})
    # по каждому веществу оставляем максимум
    best: dict[str, dict] = {}
    for r in rows:
        cur = best.get(r["code"])
        if cur is None or r["max_pdk"] > cur["max_pdk"]:
            best[r["code"]] = r
    return sorted(best.values(), key=lambda r: -r["max_pdk"])


def import_uprza_results(project: str, path: Path) -> dict[str, Any]:
    rows = parse_uprza_file(path)
    if not rows:
        # НЕ затираем прежний корректный импорт пустым результатом (ревью)
        raise RuntimeError(
            "не распознано ни одной строки «код ЗВ + доли ПДК» — похоже, формат "
            "выгрузки другой. Прежний импорт (если был) сохранён. Пришлите файл "
            "разработчику — парсер расширяется быстро.")
    data = {
        "file": path.name,
        "imported_at": datetime.now().isoformat(timespec="seconds"),
        "rows": rows,
        "exceedances": [r for r in rows if r["max_pdk"] > 1.0],
    }
    out = results_path(project)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return data


def load_uprza_results(project: str) -> dict[str, Any] | None:
    p = results_path(project)
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:  # noqa: BLE001
        return None
